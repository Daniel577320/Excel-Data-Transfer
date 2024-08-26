import tkinter as tk
from tkinter import filedialog, messagebox
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment
import shutil
from datetime import datetime
import os
import sys

def resource_path(relative_path):
    try:
        base_path = sys._MEIPASS
    except Exception:
        base_path = os.path.abspath(".")
    return os.path.join(base_path, relative_path)

def select_file(file_path_var):
    filepath = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx *.xls")])
    if filepath:
        file_path_var.set(filepath)

def transfer_data():
    try:
        source_file = source_path.get()
        sheet_name = 'BESTUURDERS CLIPBOARD'
        destination_file = resource_path('BESTUURDERS WERKSTAAT.xlsx')

        if not source_file:
            raise ValueError("Source file must be selected")

        # Read the source file into a DataFrame
        source_df = pd.read_excel(source_file, sheet_name=sheet_name, header=None)
        source_df.columns = source_df.iloc[2]

        # Extracting date
        date_cell_value = str(source_df.iloc[1, 0])
        date_obj = datetime.strptime(date_cell_value, "%Y-%m-%d %H:%M:%S")
        day_of_week = date_obj.strftime("%A")
        formatted_date = date_obj.strftime("%d %B %Y")

        # Extract the data from the source DataFrame
        data = source_df.iloc[3:, 1:13]

        #Get it in correct format
        data["T P/UP"] = data["T P/UP"].astype(str)

        data = data.sort_values(by=["DRIVER", "T P/UP"])
        print(data)

        # Load the existing workbook
        workbook = load_workbook(destination_file)
        original_sheet = workbook[sheet_name]

        # Track sheet creation for unique pairs
        drivers_written = 0

        first_driver_row = 5
        second_driver_row = 15
        first_driver_cell = "A2"
        second_driver_cell = "A12"

        # Initialize a variable to keep track of the current sheet
        current_sheet = original_sheet

        for group_name, group_data in data.groupby("DRIVER"):
            num_trips = len(group_data)  # Get the number of trips for the current driver
        
          
            if drivers_written % 2 == 0:
                if drivers_written != 0:
                    # Create a new sheet for every pair of drivers
                    current_sheet = workbook.copy_worksheet(original_sheet)
                    new_sheet_title = f"Drivers {drivers_written + 1}-{drivers_written + 2}"
                    current_sheet.title = new_sheet_title
                    
                # Clear previous data for the new sheet
                current_sheet[first_driver_cell].value = ""
                current_sheet[second_driver_cell].value = ""

            # Determine where to start writing based on whether it's the first or second driver
            current_cell = first_driver_cell if drivers_written % 2 == 0 else second_driver_cell
            date_cell = "A3" if drivers_written % 2 == 0 else "A13"
            start_row = first_driver_row if drivers_written % 2 == 0 else second_driver_row
            max_rows = 5  # The maximum number of rows to write per driver

            # Clear all rows before writing new data to avoid leftover data
            for row in range(start_row, start_row + max_rows):
                for col in range(2, 8):  # Assuming columns 2 to 7 are used for trip data
                    current_sheet.cell(row=row, column=col).value = None

            # Write header for the driver
            current_sheet[current_cell].value = f"BESTUURDER:      {group_name}"
            current_sheet[date_cell].value = f"DAG:  {day_of_week}                      DATUM:  {formatted_date}"

          

            # Write only the number of trips for this driver
            for idx, row in group_data.iterrows():
                for col, value in enumerate([row[3], row[5], row[6], row[7], row[8], row[11]], start=2):
                    cell = current_sheet.cell(row=start_row, column=col, value=value)
                    cell.font = Font(size=20)
                    cell.alignment = Alignment(horizontal='center')
                start_row += 1

            drivers_written += 1

        # Save the modified workbook back to the destination file
        workbook.save(destination_file)
        messagebox.showinfo("Success", "Data transferred successfully")

    except Exception as e:
        messagebox.showerror("Error", str(e))



def reset_file():
    try:
        destination_file = resource_path('BESTUURDERS WERKSTAAT.xlsx')
        backup_file = resource_path('BESTUURDERS WERKSTAAT - backup.xlsx')

        # Copy the backup file to the destination file, overwriting it
        shutil.copyfile(backup_file, destination_file)

        messagebox.showinfo("Success", "Excel file reset successfully")

    except Exception as e:
        messagebox.showerror("Error", str(e))

def view_files():
    try:
        # Assuming the modified Excel files are saved in the same directory as the executable
        directory = os.path.dirname(os.path.abspath(__file__))
        excel_file_path = os.path.join(directory, 'BESTUURDERS WERKSTAAT.xlsx')

        # Open the modified Excel files directly in Excel
        os.startfile(excel_file_path)

    except Exception as e:
        messagebox.showerror("Error", str(e))

# Set up the GUI
root = tk.Tk()
root.title("Excel Data Transfer")

source_path = tk.StringVar()

tk.Label(root, text="Select Source Excel File:").grid(row=0, column=0, padx=10, pady=10)
tk.Entry(root, textvariable=source_path, width=50).grid(row=0, column=1, padx=10, pady=10)
tk.Button(root, text="Browse", command=lambda: select_file(source_path)).grid(row=0, column=2, padx=10, pady=10)

tk.Button(root, text="Transfer Data", command=transfer_data).grid(row=1, column=0, padx=5, pady=5)
tk.Button(root, text="View Excel File", command=view_files).grid(row=1, column=2, padx=5, pady=5)
tk.Button(root, text="Reset Excel File", command=reset_file).grid(row=1, column=1, padx=5, pady=5)

root.mainloop()